from flask import Flask
import os
import openpyxl

# using templates
from flask import render_template, make_response, url_for, send_file, request

from flask_cors import CORS, cross_origin

application = Flask(__name__)
cors = CORS(application, resources={r"/*": {"origins": "*"}})
# app.config['CORS_HEADERS'] = 'Content-Type'


@application.route("/")
@application.route("/<name>")
@cross_origin()
def index(name=None):
    return "Hello my friend"


@application.route("/get_excel")
@cross_origin()
def get_excel():
    res = make_response(
        send_file(
            os.path.join("static", "excel", "Partner-Hosted-Self-assessment.xlsx")
        )
    )
    res.headers["Content-Type"] = "blob"

    print("Response!!!!", res)

    return res


@application.route("/gen_excel", methods=["GET", "POST"])
@cross_origin()
def gen_excel():
    wb = openpyxl.load_workbook(
        os.path.join("static", "excel", "FTR-self-assessment.xlsx")
    )
    ws = wb["Partner Hosted FTR Requirements"]
    controls = request.get_json()
    for control in controls["body"]:
        ws[control["comment_cell"]] = control["comment"]
        ws[control["value_cell"]] = control["value"]
    wb.save(os.path.join("static", "excel", "Partner-Hosted-Self-assessment.xlsx"))
    return "Done!"


if __name__ == "__main__":
    application.debug = True
    application.run()
